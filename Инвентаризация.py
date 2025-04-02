import os
import sys
import sqlite3
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                            QTableWidget, QTableWidgetItem, QDialog, QLabel, QLineEdit, QComboBox, 
                            QMessageBox, QFileDialog, QAction, QStyleFactory, QAbstractItemView, QTabWidget,
                            QDialogButtonBox, QFormLayout, QCheckBox, QGroupBox, QFrame, QHeaderView,
                            QMenu, QStatusBar)
from PyQt5.QtCore import Qt, QSettings, QSize
from PyQt5.QtGui import QIcon, QFont, QColor, QPalette
from PyQt5.QtWidgets import QStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtGui import QPainter
from PyQt5.QtWidgets import QInputDialog
from PyQt5.QtWidgets import QSizePolicy
from PyQt5.QtWidgets import QHeaderView

class AssetDialog(QDialog):
    def __init__(self, parent=None, asset_data=None, mode='add'):
        super().__init__(parent)
        self.mode = mode
        self.setWindowTitle("Добавить ОС" if mode == 'add' else "Редактировать ОС")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        self.setMinimumSize(500, 600)
        
        layout = QVBoxLayout()
        
        # Form layout
        form_layout = QFormLayout()
        form_layout.setVerticalSpacing(15)
        
        # Inventory Number
        self.inventory_number = QLineEdit()
        form_layout.addRow("Инв. №:", self.inventory_number)
        
        # Asset Type with add/remove buttons
        asset_type_layout = QHBoxLayout()
        self.asset_type = QComboBox()
        self.asset_type.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        asset_type_layout.addWidget(self.asset_type)
        
        self.add_asset_type_btn = QPushButton("+")
        self.add_asset_type_btn.setFixedWidth(30)
        self.add_asset_type_btn.clicked.connect(self.add_new_asset_type)
        asset_type_layout.addWidget(self.add_asset_type_btn)
        
        self.remove_asset_type_btn = QPushButton("−")
        self.remove_asset_type_btn.setFixedWidth(30)
        self.remove_asset_type_btn.clicked.connect(self.remove_asset_type)
        asset_type_layout.addWidget(self.remove_asset_type_btn)
        
        form_layout.addRow("Тип ОС:", asset_type_layout)
        
        # Name
        self.name = QLineEdit()
        form_layout.addRow("Наименование:", self.name)
        
        # Serial Number
        self.serial_number = QLineEdit()
        form_layout.addRow("Серийный №:", self.serial_number)
        
        # Status
        self.status = QComboBox()
        self.status.addItems(["В работе", "На складе", "Требуется ремонт", "Вышел из строя"])
        form_layout.addRow("Статус:", self.status)
        
        # Department with add/remove buttons
        department_layout = QHBoxLayout()
        self.department = QComboBox()
        self.department.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        department_layout.addWidget(self.department)
        
        self.add_department_btn = QPushButton("+")
        self.add_department_btn.setFixedWidth(30)
        self.add_department_btn.clicked.connect(self.add_new_department)
        department_layout.addWidget(self.add_department_btn)
        
        self.remove_department_btn = QPushButton("−")
        self.remove_department_btn.setFixedWidth(30)
        self.remove_department_btn.clicked.connect(self.remove_department)
        department_layout.addWidget(self.remove_department_btn)
        
        form_layout.addRow("Подразделение:", department_layout)
        
        # MOL (Materially Responsible Person)
        self.mol = QLineEdit()
        form_layout.addRow("МОЛ:", self.mol)
        
        # Location
        self.location = QLineEdit()
        form_layout.addRow("Местоположение:", self.location)
        
        # Purchase Date
        self.purchase_date = QLineEdit()
        self.purchase_date.setPlaceholderText("дд.мм.гггг")
        form_layout.addRow("Дата постановки на учет:", self.purchase_date)
        
        # Notes
        self.notes = QLineEdit()
        form_layout.addRow("Примечания:", self.notes)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        # Load data from database if parent has db_manager
        if hasattr(parent, 'db_manager') and parent.db_manager:
            self.load_reference_data()
        
        if asset_data:
            self.fill_form(asset_data)
            
        self.update_delete_buttons_state()
        self.asset_type.currentIndexChanged.connect(self.update_delete_buttons_state)
        self.department.currentIndexChanged.connect(self.update_delete_buttons_state)
        
        if mode == 'view':
            self.history_btn = QPushButton("История изменений")
            self.history_btn.clicked.connect(self.show_history)
            button_box.addButton(self.history_btn, QDialogButtonBox.ActionRole)

    def show_history(self):
        asset_id = self.asset_data.get('id') if hasattr(self, 'asset_data') else None
        if asset_id and hasattr(self.parent(), 'db_manager'):
            dialog = HistoryDialog(self, asset_id, self.parent().db_manager)
            dialog.exec_()
        
    def load_reference_data(self):
        """Load asset types and departments from database"""
        db = self.parent().db_manager
        if not db:
            return
            
        # Load asset types
        self.asset_type.clear()
        types = db.get_asset_types()
        for item in types:
            self.asset_type.addItem(item['name'], item['id'])
        
        # Load departments
        self.department.clear()
        depts = db.get_departments()
        for item in depts:
            self.department.addItem(item['name'], item['id'])
    
    def add_new_asset_type(self):
        """Add new asset type to database"""
        if not hasattr(self.parent(), 'db_manager') or not self.parent().db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        name, ok = QInputDialog.getText(
            self, 
            "Добавить тип ОС", 
            "Введите название нового типа ОС:",
            QLineEdit.Normal
        )
        
        if ok and name.strip():
            try:
                db = self.parent().db_manager
                type_id = db.add_asset_type(name.strip())
                
                # Refresh combobox
                self.load_reference_data()
                
                # Select newly added item
                index = self.asset_type.findData(type_id)
                if index >= 0:
                    self.asset_type.setCurrentIndex(index)
                    
                QMessageBox.information(self, "Успех", "Тип ОС успешно добавлен")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить тип ОС: {str(e)}")
    
    def add_new_department(self):
        """Add new department to database"""
        if not hasattr(self.parent(), 'db_manager') or not self.parent().db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        name, ok = QInputDialog.getText(
            self, 
            "Добавить подразделение", 
            "Введите название нового подразделения:",
            QLineEdit.Normal
        )
        
        if ok and name.strip():
            try:
                db = self.parent().db_manager
                dept_id = db.add_department(name.strip())
                
                # Refresh combobox
                self.load_reference_data()
                
                # Select newly added item
                index = self.department.findData(dept_id)
                if index >= 0:
                    self.department.setCurrentIndex(index)
                    
                QMessageBox.information(self, "Успех", "Подразделение успешно добавлено")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить подразделение: {str(e)}")
    
    def fill_form(self, data):
        self.inventory_number.setText(data.get('inventory_number', ''))
        self.name.setText(data.get('name', ''))
        self.serial_number.setText(data.get('serial_number', ''))
        self.mol.setText(data.get('mol', ''))
        self.location.setText(data.get('location', ''))
        self.notes.setText(data.get('notes', ''))
        
        # For backward compatibility
        if 'asset_type_id' in data:
            index = self.asset_type.findData(data['asset_type_id'])
            if index >= 0:
                self.asset_type.setCurrentIndex(index)
        elif 'asset_type' in data:
            index = self.asset_type.findText(data.get('asset_type', ''))
            if index >= 0:
                self.asset_type.setCurrentIndex(index)
                
        if 'department_id' in data:
            index = self.department.findData(data['department_id'])
            if index >= 0:
                self.department.setCurrentIndex(index)
        elif 'department' in data:
            index = self.department.findText(data.get('department', ''))
            if index >= 0:
                self.department.setCurrentIndex(index)
                
        purchase_date = data.get('purchase_date', '')
        if purchase_date:
            try:
                date_obj = datetime.strptime(purchase_date, "%Y-%m-%d")
                self.purchase_date.setText(date_obj.strftime("%d.%m.%Y"))
            except ValueError:
                self.purchase_date.setText(purchase_date)
    
    def validate_and_accept(self):
        if not self.inventory_number.text().strip():
            QMessageBox.warning(self, "Ошибка", "Поле 'Инв. №' обязательно для заполнения")
            return
            
        if not self.name.text().strip():
            QMessageBox.warning(self, "Ошибка", "Поле 'Наименование' обязательно для заполнения")
            return
            
        purchase_date = self.purchase_date.text().strip()
        if purchase_date:
            try:
                datetime.strptime(purchase_date, "%d.%m.%Y")
            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Дата должна быть в формате дд.мм.гггг")
                return
                
        self.accept()
    
    def get_data(self):
        purchase_date = self.purchase_date.text().strip()
        if purchase_date:
            try:
                date_obj = datetime.strptime(purchase_date, "%d.%m.%Y")
                purchase_date = date_obj.strftime("%Y-%m-%d")
            except ValueError:
                purchase_date = ''
        
        return {
            'inventory_number': self.inventory_number.text().strip(),
            'asset_type_id': self.asset_type.currentData(),
            'name': self.name.text().strip(),
            'serial_number': self.serial_number.text().strip(),
            'status': self.status.currentText(),
            'department_id': self.department.currentData(),
            'mol': self.mol.text().strip(),
            'location': self.location.text().strip(),
            'purchase_date': purchase_date,
            'notes': self.notes.text().strip()
        }
        self.update_delete_buttons_state()
        self.asset_type.currentIndexChanged.connect(self.update_delete_buttons_state)
        self.department.currentIndexChanged.connect(self.update_delete_buttons_state)

    def update_delete_buttons_state(self):
        """Обновляет состояние кнопок удаления"""
        self.remove_asset_type_btn.setEnabled(self.asset_type.currentIndex() >= 0)
        self.remove_department_btn.setEnabled(self.department.currentIndex() >= 0)

    def remove_asset_type(self):
        """Удаляет выбранный тип ОС"""
        if not hasattr(self.parent(), 'db_manager') or not self.parent().db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        current_id = self.asset_type.currentData()
        current_text = self.asset_type.currentText()
        
        if not current_id:
            return
            
        reply = QMessageBox.question(
            self,
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить тип ОС '{current_text}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                db = self.parent().db_manager
                db.delete_asset_type(current_id)
                
                # Обновляем список типов ОС
                self.load_reference_data()
                QMessageBox.information(self, "Успех", "Тип ОС успешно удален")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось удалить тип ОС: {str(e)}")

    def remove_department(self):
        """Удаляет выбранное подразделение"""
        if not hasattr(self.parent(), 'db_manager') or not self.parent().db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        current_id = self.department.currentData()
        current_text = self.department.currentText()
        
        if not current_id:
            return
            
        reply = QMessageBox.question(
            self,
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить подразделение '{current_text}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                db = self.parent().db_manager
                db.delete_department(current_id)
                
                # Обновляем список подразделений
                self.load_reference_data()
                QMessageBox.information(self, "Успех", "Подразделение успешно удалено")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось удалить подразделение: {str(e)}")


class FilterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Фильтр")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setMinimumSize(500, 600)
        
        layout = QVBoxLayout()
        
        # Create tab widget
        tab_widget = QTabWidget()
        
        # Main filters tab
        main_tab = QWidget()
        main_layout = QFormLayout()
        main_layout.setVerticalSpacing(15)
        
        # Inventory Number
        self.inventory_number_filter = QLineEdit()
        main_layout.addRow("Инв. №:", self.inventory_number_filter)
        
        # Asset Type
        self.asset_type_filter = QComboBox()
        self.asset_type_filter.addItem("(Все)", "")
        # Загружаем типы ОС из базы данных, если есть подключение
        if hasattr(parent, 'db_manager') and parent.db_manager:
            types = parent.db_manager.get_asset_types()
            for item in types:
                self.asset_type_filter.addItem(item['name'], item['id'])
        else:
            # Fallback статический список
            self.asset_type_filter.addItems(["Принтер", "Системный блок", "Ноутбук", "Монитор", "ИБП"])
        
        main_layout.addRow("Тип ОС:", self.asset_type_filter)
        
        # Name
        self.name_filter = QLineEdit()
        main_layout.addRow("Наименование:", self.name_filter)
        
        # Serial Number
        self.serial_number_filter = QLineEdit()
        main_layout.addRow("Серийный №:", self.serial_number_filter)
        
        # Status
        self.status_filter = QComboBox()
        self.status_filter.addItem("(Все)", "")
        self.status_filter.addItems(["В работе", "На складе", "Требуется ремонт", "Вышел из строя"])
        main_layout.addRow("Статус:", self.status_filter)
        
        # Department
        self.department_filter = QComboBox()
        self.department_filter.addItem("(Все)", "")
        # Загружаем подразделения из базы данных, если есть подключение
        if hasattr(parent, 'db_manager') and parent.db_manager:
            depts = parent.db_manager.get_departments()
            for item in depts:
                self.department_filter.addItem(item['name'], item['id'])
        else:
            # Fallback статический список
            departments = [
                "МТО", "ОКС", "СБ", "СГИ", "СГЭ", "Служба обеспечения производства", 
                "СОТиПБ", "СпГР", "Служба по металлургии", "Служба по минеральным ресурсам", 
                "Служба технического обслуживания и ремонтов", "Служба управления персоналом", 
                "Финансово-экономическая служба", "Администрация", "Бизнес-система", 
                "Отдел охраны окружающей среды", "Отдел технического контроля"
            ]
            self.department_filter.addItems(departments)
        
        main_layout.addRow("Подразделение:", self.department_filter)
        
        # MOL
        self.mol_filter = QLineEdit()
        main_layout.addRow("МОЛ:", self.mol_filter)
        
        # Location
        self.location_filter = QLineEdit()
        main_layout.addRow("Местоположение:", self.location_filter)
        
        main_tab.setLayout(main_layout)
        tab_widget.addTab(main_tab, "Основные фильтры")
        
        # Date filters tab
        date_tab = QWidget()
        date_layout = QFormLayout()
        date_layout.setVerticalSpacing(15)
        
        # Purchase Date From
        self.purchase_date_from = QLineEdit()
        self.purchase_date_from.setPlaceholderText("дд.мм.гггг")
        date_layout.addRow("Дата постановки на учет (с):", self.purchase_date_from)
        
        # Purchase Date To
        self.purchase_date_to = QLineEdit()
        self.purchase_date_to.setPlaceholderText("дд.мм.гггг")
        date_layout.addRow("Дата постановки на учет (по):", self.purchase_date_to)
        
        date_tab.setLayout(date_layout)
        tab_widget.addTab(date_tab, "Фильтры по дате")
        
        layout.addWidget(tab_widget)
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel | QDialogButtonBox.Reset
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        reset_btn = button_box.button(QDialogButtonBox.Reset)
        reset_btn.clicked.connect(self.reset_filters)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
    
    def reset_filters(self):
        self.inventory_number_filter.clear()
        self.asset_type_filter.setCurrentIndex(0)
        self.name_filter.clear()
        self.serial_number_filter.clear()
        self.status_filter.setCurrentIndex(0)
        self.department_filter.setCurrentIndex(0)
        self.mol_filter.clear()
        self.location_filter.clear()
        self.purchase_date_from.clear()
        self.purchase_date_to.clear()
    
    def get_filters(self):
        filters = {}
        
        if self.inventory_number_filter.text().strip():
            filters['inventory_number'] = self.inventory_number_filter.text().strip()
            
        if self.asset_type_filter.currentIndex() > 0:
            # Используем ID типа ОС вместо текста
            filters['asset_type_id'] = self.asset_type_filter.currentData()
            
        if self.name_filter.text().strip():
            filters['name'] = self.name_filter.text().strip()
            
        if self.serial_number_filter.text().strip():
            filters['serial_number'] = self.serial_number_filter.text().strip()
            
        if self.status_filter.currentIndex() > 0:
            filters['status'] = self.status_filter.currentText()
            
        if self.department_filter.currentIndex() > 0:
            # Используем ID подразделения вместо текста
            filters['department_id'] = self.department_filter.currentData()
            
        if self.mol_filter.text().strip():
            filters['mol'] = self.mol_filter.text().strip()
            
        if self.location_filter.text().strip():
            filters['location'] = self.location_filter.text().strip()
            
        # Date filters
        date_from = self.purchase_date_from.text().strip()
        date_to = self.purchase_date_to.text().strip()
        
        if date_from:
            try:
                datetime.strptime(date_from, "%d.%m.%Y")
                filters['purchase_date_from'] = date_from
            except ValueError:
                pass
                
        if date_to:
            try:
                datetime.strptime(date_to, "%d.%m.%Y")
                filters['purchase_date_to'] = date_to
            except ValueError:
                pass
                
        return filters


class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None
        self.connect()
        self.create_tables()
    
    def connect(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.execute("PRAGMA foreign_keys = ON")
    
    def create_tables(self):
        cursor = self.conn.cursor()
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS asset_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT
        )
        """)
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT
        )
        """)
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inventory_number TEXT UNIQUE NOT NULL,
            asset_type_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            serial_number TEXT,
            status TEXT NOT NULL,
            department_id INTEGER NOT NULL,
            mol TEXT,
            location TEXT,
            purchase_date TEXT,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (asset_type_id) REFERENCES asset_types(id),
            FOREIGN KEY (department_id) REFERENCES departments(id)
        )
        """)
        
        # Таблица истории (оставить без изменений)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS asset_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER NOT NULL,
            changed_field TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            changed_at TEXT DEFAULT CURRENT_TIMESTAMP,
            changed_by TEXT,
            FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
        )
        """)
        
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_asset_history_asset_id ON asset_history(asset_id)")
        
        self._initialize_reference_data()
        self.conn.commit()
        
    def log_asset_change(self, asset_id, changed_field, old_value, new_value, changed_by=None):
        cursor = self.conn.cursor()
        cursor.execute(
            "INSERT INTO asset_history (asset_id, changed_field, old_value, new_value, changed_by) VALUES (?, ?, ?, ?, ?)",
            (asset_id, changed_field, old_value, new_value, changed_by)
        )
        self.conn.commit()
        
        self.conn.commit()
        
        self._initialize_reference_data()
        
        self.conn.commit()
        
    def _initialize_reference_data(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM asset_types")
        if cursor.fetchone()[0] == 0:
            default_types = ["Принтер", "Системный блок", "Ноутбук", "Монитор", "ИБП"]
            for type_name in default_types:
                cursor.execute("INSERT INTO asset_types (name) VALUES (?)", (type_name,))
        cursor.execute("SELECT COUNT(*) FROM departments")
        if cursor.fetchone()[0] == 0:
            default_departments = [
                "МТО", "ОКС", "СБ", "СГИ", "СГЭ", "Служба обеспечения производства", 
                "СОТиПБ", "СпГР", "Служба по металлургии", "Служба по минеральным ресурсам", 
                "Служба технического обслуживания и ремонтов", "Служба управления персоналом", 
                "Финансово-экономическая служба", "Администрация", "Бизнес-система", 
                "Отдел охраны окружающей среды", "Отдел технического контроля"
            ]
            for dept_name in default_departments:
                cursor.execute("INSERT INTO departments (name) VALUES (?)", (dept_name,))
        
        self.conn.commit()
        
        # Методы для работы с типами ОС
    def get_asset_types(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, name, description FROM asset_types ORDER BY name")
        return [{'id': row[0], 'name': row[1], 'description': row[2]} for row in cursor.fetchall()]
    
    def add_asset_type(self, name, description=None):
        cursor = self.conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO asset_types (name, description) VALUES (?, ?)",
                (name, description)
            )
            self.conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            raise ValueError("Тип ОС с таким названием уже существует")
    
    def update_asset_type(self, type_id, name, description=None):
        cursor = self.conn.cursor()
        cursor.execute(
            "UPDATE asset_types SET name = ?, description = ? WHERE id = ?",
            (name, description, type_id)
        )
        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_asset_type(self, type_id):
        cursor = self.conn.cursor()
        try:
            cursor.execute("DELETE FROM asset_types WHERE id = ?", (type_id,))
            self.conn.commit()
            return cursor.rowcount > 0
        except sqlite3.IntegrityError:
            raise ValueError("Невозможно удалить тип ОС, так как он используется в записях активов")
    def get_departments(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, name, description FROM departments ORDER BY name")
        return [{'id': row[0], 'name': row[1], 'description': row[2]} for row in cursor.fetchall()]
    
    def add_department(self, name, description=None):
        cursor = self.conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO departments (name, description) VALUES (?, ?)",
                (name, description)
            )
            self.conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            raise ValueError("Подразделение с таким названием уже существует")
    
    def update_department(self, dept_id, name, description=None):
        cursor = self.conn.cursor()
        cursor.execute(
            "UPDATE departments SET name = ?, description = ? WHERE id = ?",
            (name, description, dept_id)
        )
        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_department(self, dept_id):
        cursor = self.conn.cursor()
        try:
            cursor.execute("DELETE FROM departments WHERE id = ?", (dept_id,))
            self.conn.commit()
            return cursor.rowcount > 0
        except sqlite3.IntegrityError:
            raise ValueError("Невозможно удалить подразделение, так как оно используется в записях активов")
    
    def add_asset(self, asset_data):
        cursor = self.conn.cursor()
        
        query = """
        INSERT INTO assets (
            inventory_number, asset_type_id, name, serial_number, status, 
            department_id, mol, location, purchase_date, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        params = (
            asset_data['inventory_number'],
            asset_data['asset_type_id'],  # Используем ID вместо текста
            asset_data['name'],
            asset_data['serial_number'],
            asset_data['status'],
            asset_data['department_id'],  # Используем ID вместо текста
            asset_data['mol'],
            asset_data['location'],
            asset_data['purchase_date'],
            asset_data['notes']
        )
        
        try:
            cursor.execute(query, params)
            self.conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            raise ValueError("Основное средство с таким инвентарным номером уже существует")
            
    def get_assets(self, filters=None):
        cursor = self.conn.cursor()
        
        query = """
        SELECT a.*, at.name as asset_type, d.name as department 
        FROM assets a
        LEFT JOIN asset_types at ON a.asset_type_id = at.id
        LEFT JOIN departments d ON a.department_id = d.id
        """
        params = []
        
        if filters:
            conditions = []
            
            if 'inventory_number' in filters:
                conditions.append("a.inventory_number LIKE ?")
                params.append(f"%{filters['inventory_number']}%")
                
            if 'asset_type_id' in filters:  # Изменено с 'asset_type' на 'asset_type_id'
                conditions.append("a.asset_type_id = ?")
                params.append(filters['asset_type_id'])
                
            if 'name' in filters:
                conditions.append("a.name LIKE ?")
                params.append(f"%{filters['name']}%")
                
            if 'serial_number' in filters:
                conditions.append("a.serial_number LIKE ?")
                params.append(f"%{filters['serial_number']}%")
                
            if 'status' in filters:
                conditions.append("a.status = ?")
                params.append(filters['status'])
                
            if 'department_id' in filters:  # Изменено с 'department' на 'department_id'
                conditions.append("a.department_id = ?")
                params.append(filters['department_id'])
                
            if 'mol' in filters:
                conditions.append("a.mol LIKE ?")
                params.append(f"%{filters['mol']}%")
                
            if 'location' in filters:
                conditions.append("a.location LIKE ?")
                params.append(f"%{filters['location']}%")
                
            # Date filters
            if 'purchase_date_from' in filters:
                try:
                    date_obj = datetime.strptime(filters['purchase_date_from'], "%d.%m.%Y")
                    conditions.append("a.purchase_date >= ?")
                    params.append(date_obj.strftime("%Y-%m-%d"))
                except ValueError:
                    pass
                    
            if 'purchase_date_to' in filters:
                try:
                    date_obj = datetime.strptime(filters['purchase_date_to'], "%d.%m.%Y")
                    conditions.append("a.purchase_date <= ?")
                    params.append(date_obj.strftime("%Y-%m-%d"))
                except ValueError:
                    pass
                    
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY a.inventory_number"
        
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        results = cursor.fetchall()
        
        assets = []
        for row in results:
            asset = dict(zip(columns, row))
            if asset.get('purchase_date'):
                try:
                    date_obj = datetime.strptime(asset['purchase_date'], "%Y-%m-%d")
                    asset['purchase_date'] = date_obj.strftime("%d.%m.%Y")
                except ValueError:
                    pass
            assets.append(asset)
        
        return assets
        
    def get_asset(self, asset_id):
        """Получить данные одного актива по ID"""
        cursor = self.conn.cursor()
        cursor.execute("""
        SELECT a.*, at.name as asset_type, d.name as department 
        FROM assets a
        LEFT JOIN asset_types at ON a.asset_type_id = at.id
        LEFT JOIN departments d ON a.department_id = d.id
        WHERE a.id = ?
        """, (asset_id,))
        
        result = cursor.fetchone()
        if not result:
            return None
            
        columns = [column[0] for column in cursor.description]
        asset = dict(zip(columns, result))
        
        if asset.get('purchase_date'):
            try:
                date_obj = datetime.strptime(asset['purchase_date'], "%Y-%m-%d")
                asset['purchase_date'] = date_obj.strftime("%d.%m.%Y")
            except ValueError:
                pass
                
        return asset
    
    def update_asset(self, asset_id, asset_data, changed_by=None):
        # Получаем старые данные
        old_data = self.get_asset(asset_id)
        if not old_data:
            return False

        cursor = self.conn.cursor()
        
        # Обновляем запись
        cursor.execute("""
        UPDATE assets SET
            inventory_number = ?,
            asset_type_id = ?,
            name = ?,
            serial_number = ?,
            status = ?,
            department_id = ?,
            mol = ?,
            location = ?,
            purchase_date = ?,
            notes = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """, (
            asset_data['inventory_number'],
            asset_data['asset_type_id'],
            asset_data['name'],
            asset_data['serial_number'],
            asset_data['status'],
            asset_data['department_id'],
            asset_data['mol'],
            asset_data['location'],
            asset_data['purchase_date'],
            asset_data['notes'],
            asset_id
        ))

        # Логируем изменения для каждого поля
        fields_to_check = [
            'inventory_number', 'asset_type_id', 'name', 'serial_number',
            'status', 'department_id', 'mol', 'location', 'purchase_date', 'notes'
        ]
        
        for field in fields_to_check:
            old_val = str(old_data.get(field, ''))
            new_val = str(asset_data.get(field, ''))
            if old_val != new_val:
                self.log_asset_change(asset_id, field, old_val, new_val, changed_by)

        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_asset(self, asset_id):
        """Удаление актива по ID"""
        cursor = self.conn.cursor()
        try:
            # Просто удаляем запись - каскадное удаление настроено в FOREIGN KEY
            cursor.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
            self.conn.commit()
            return cursor.rowcount > 0
        except sqlite3.Error as e:
            self.conn.rollback()
            raise Exception(f"Не удалось удалить основное средство: {str(e)}")

        # Логируем изменения для каждого поля
        fields_to_check = [
            'inventory_number', 'asset_type_id', 'name', 'serial_number',
            'status', 'department_id', 'mol', 'location', 'purchase_date', 'notes'
        ]
        
        for field in fields_to_check:
            old_val = str(old_data.get(field, ''))
            new_val = str(asset_data.get(field, ''))
            if old_val != new_val:
                self.log_asset_change(asset_id, field, old_val, new_val, changed_by)

        self.conn.commit()
        return cursor.rowcount > 0
        
    def get_asset_history(self, asset_id):
        cursor = self.conn.cursor()
        cursor.execute("""
        SELECT 
            changed_field, 
            old_value, 
            new_value, 
            datetime(changed_at, 'localtime') as changed_at,
            changed_by
        FROM asset_history
        WHERE asset_id = ?
        ORDER BY changed_at DESC
        """, (asset_id,))
        
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    def close(self):
        if self.conn:
            self.conn.close()
            
    def delete_asset_type(self, type_id):
        cursor = self.conn.cursor()
        try:
            cursor.execute("DELETE FROM asset_types WHERE id = ?", (type_id,))
            self.conn.commit()
            return cursor.rowcount > 0
        except sqlite3.IntegrityError:
            raise ValueError("Невозможно удалить тип ОС, так как он используется в записях активов")
    
    def delete_department(self, dept_id):
        cursor = self.conn.cursor()
        try:
            cursor.execute("DELETE FROM departments WHERE id = ?", (dept_id,))
            self.conn.commit()
            return cursor.rowcount > 0
        except sqlite3.IntegrityError:
            raise ValueError("Невозможно удалить подразделение, так как оно используется в записях активов")
            
class ReferenceItemDialog(QDialog):
    def __init__(self, parent=None, item_type='asset_type', item_data=None, mode='add'):
        super().__init__(parent)
        self.mode = mode
        self.item_type = item_type
        self.setWindowTitle(f"{'Добавить' if mode == 'add' else 'Редактировать'} {'тип ОС' if item_type == 'asset_type' else 'подразделение'}")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        layout = QVBoxLayout()
        
        # Form layout
        form_layout = QFormLayout()
        form_layout.setVerticalSpacing(15)
        
        # Name
        self.name_edit = QLineEdit()
        form_layout.addRow("Название:", self.name_edit)
        
        # Description
        self.description_edit = QLineEdit()
        form_layout.addRow("Описание:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        if item_data:
            self.fill_form(item_data)
    
    def fill_form(self, data):
        self.name_edit.setText(data.get('name', ''))
        self.description_edit.setText(data.get('description', ''))
    
    def validate_and_accept(self):
        if not self.name_edit.text().strip():
            item_name = "тип ОС" if self.item_type == 'asset_type' else "подразделение"
            QMessageBox.warning(self, "Ошибка", f"Поле 'Название' обязательно для заполнения {item_name}")
            return
        self.accept()
    
    def get_data(self):
        return {
            'name': self.name_edit.text().strip(),
            'description': self.description_edit.text().strip()
        }

class ReferenceManagerDialog(QDialog):
    def __init__(self, parent=None, item_type='asset_type'):
        super().__init__(parent)
        self.item_type = item_type
        self.db_manager = parent.db_manager if parent and hasattr(parent, 'db_manager') else None
        self.setWindowTitle(f"Управление {'типами ОС' if item_type == 'asset_type' else 'подразделениями'}")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Название", "Описание"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("Добавить")
        self.add_btn.clicked.connect(self.add_item)
        
        self.edit_btn = QPushButton("Редактировать")
        self.edit_btn.clicked.connect(self.edit_item)
        
        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.clicked.connect(self.delete_item)
        
        button_layout.addWidget(self.add_btn)
        button_layout.addWidget(self.edit_btn)
        button_layout.addWidget(self.delete_btn)
        
        # Main layout
        layout.addWidget(self.table)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        self.load_data()
        self.update_buttons_state()
        self.table.selectionModel().selectionChanged.connect(self.update_buttons_state)
    
    def load_data(self):
        if not self.db_manager:
            return
            
        # Получаем данные из базы
        if self.item_type == 'asset_type':
            items = self.db_manager.get_asset_types()
        else:
            items = self.db_manager.get_departments()
        
        # Устанавливаем количество колонок
        self.table.setColumnCount(2)  # Название и Описание
        
        # Устанавливаем заголовки
        self.table.setHorizontalHeaderLabels(["Название", "Описание"])
        
        self.table.setRowCount(len(items))
        
        for row, item in enumerate(items):
            self.table.setItem(row, 0, QTableWidgetItem(item['name']))
            # Добавляем отображение описания
            description = item.get('description', '')
            self.table.setItem(row, 1, QTableWidgetItem(description))
            self.table.item(row, 0).setData(Qt.UserRole, item['id'])
    
    def update_buttons_state(self):
        has_selection = self.table.currentRow() >= 0
        self.edit_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)
    
    def add_item(self):
        dialog = ReferenceItemDialog(self, self.item_type)
        if dialog.exec_() == QDialog.Accepted:
            try:
                data = dialog.get_data()
                if self.item_type == 'asset_type':
                    self.db_manager.add_asset_type(data['name'], data['description'])
                else:
                    self.db_manager.add_department(data['name'], data['description'])
                self.load_data()
                QMessageBox.information(self, "Успех", "Запись успешно добавлена")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
    
    def edit_item(self):
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return
            
        item_id = self.table.item(selected_row, 0).data(Qt.UserRole)
        item_name = self.table.item(selected_row, 0).text()
        item_desc = self.table.item(selected_row, 1).text()
        
        dialog = ReferenceItemDialog(
            self, 
            self.item_type,
            {'name': item_name, 'description': item_desc},
            'edit'
        )
        
        if dialog.exec_() == QDialog.Accepted:
            try:
                data = dialog.get_data()
                if self.item_type == 'asset_type':
                    self.db_manager.update_asset_type(item_id, data['name'], data['description'])
                else:
                    self.db_manager.update_department(item_id, data['name'], data['description'])
                self.load_data()
                QMessageBox.information(self, "Успех", "Запись успешно обновлена")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
    
    def delete_item(self):
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return
            
        item_id = self.table.item(selected_row, 0).data(Qt.UserRole)
        item_name = self.table.item(selected_row, 0).text()
        
        reply = QMessageBox.question(
            self,
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить {'тип ОС' if self.item_type == 'asset_type' else 'подразделение'} '{item_name}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                if self.item_type == 'asset_type':
                    self.db_manager.delete_asset_type(item_id)
                else:
                    self.db_manager.delete_department(item_id)
                self.load_data()
                QMessageBox.information(self, "Успех", "Запись успешно удалена")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
                
class HistoryDialog(QDialog):
    def __init__(self, parent=None, asset_id=None, db_manager=None):
        super().__init__(parent)
        self.setWindowTitle(f"История изменений ОС #{asset_id}")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        
        # Таблица для отображения истории
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "Дата изменения", "Поле", "Старое значение", "Новое значение", "Кем изменено"
        ])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Загружаем данные
        if db_manager and asset_id:
            self.load_data(db_manager, asset_id)
        
        layout.addWidget(self.table)
        self.setLayout(layout)
    
    def load_data(self, db_manager, asset_id):
        history = db_manager.get_asset_history(asset_id)
        self.table.setRowCount(len(history))
        
        for row, record in enumerate(history):
            self.table.setItem(row, 0, QTableWidgetItem(record['changed_at']))
            self.table.setItem(row, 1, QTableWidgetItem(record['changed_field']))
            self.table.setItem(row, 2, QTableWidgetItem(record['old_value']))
            self.table.setItem(row, 3, QTableWidgetItem(record['new_value']))
            self.table.setItem(row, 4, QTableWidgetItem(record['changed_by'] or "Система"))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_menu()
        self.db_manager = None
        self.current_filters = None
        
        # Инициализируем кнопку заранее
        self.db_btn = QPushButton("Выбрать базу данных")
        self.db_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_DriveHDIcon')))
        self.db_btn.clicked.connect(self.change_database)
        
        self.setWindowTitle("Учет основных средств")
        self.setMinimumSize(1000, 600)
        
        # Settings
        self.settings = QSettings("OSWarehouse", "FixedAssetApp")
        
        # Initialize UI
        self.init_ui()
        
        # Применяем сохранённую тему
        saved_theme = self.settings.value("theme", "light")
        self.set_theme(saved_theme)
        
        # Проверяем, нужно ли показывать диалог подключения БД
        if not self.settings.value("database_path"):
            self.show_database_dialog()

    def init_ui(self):
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)
        
        # Left panel with buttons
        left_panel = QFrame()
        left_panel.setFrameShape(QFrame.StyledPanel)
        left_panel.setFixedWidth(200)
        left_layout = QVBoxLayout()
        left_panel.setLayout(left_layout)
        
        # Buttons
        self.add_btn = QPushButton("Добавить ОС")
        self.add_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileIcon')))
        self.add_btn.clicked.connect(self.add_asset)
        
        self.edit_btn = QPushButton("Редактировать ОС")
        self.edit_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileDialogDetailedView')))
        self.edit_btn.clicked.connect(self.edit_asset)
        
        self.delete_btn = QPushButton("Удалить ОС")
        self.delete_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_TrashIcon')))
        self.delete_btn.clicked.connect(self.delete_asset)
        
        self.filter_btn = QPushButton("Фильтр")
        self.filter_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileDialogContentsView')))
        self.filter_btn.clicked.connect(self.show_filter_dialog)
        
        self.clear_filter_btn = QPushButton("Сбросить фильтр")
        self.clear_filter_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_DialogResetButton')))
        self.clear_filter_btn.clicked.connect(self.clear_filters)
        
        self.print_btn = QPushButton("Печать")
        self.print_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_FileDialogListView')))
        self.print_btn.clicked.connect(self.print_assets)
        
        self.export_btn = QPushButton("Экспорт в Excel")
        self.export_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_DialogSaveButton')))
        self.export_btn.clicked.connect(self.export_to_excel)
        
        self.exit_btn = QPushButton("Выход")
        self.exit_btn.setIcon(self.style().standardIcon(getattr(QStyle, 'SP_DialogCloseButton')))
        self.exit_btn.clicked.connect(self.close)
        
        # Add buttons to left panel
        left_layout.addWidget(self.add_btn)
        left_layout.addWidget(self.edit_btn)
        left_layout.addWidget(self.delete_btn)
        left_layout.addWidget(self.filter_btn)
        left_layout.addWidget(self.clear_filter_btn)
        left_layout.addWidget(self.print_btn)
        left_layout.addWidget(self.export_btn)
        left_layout.addStretch()
        left_layout.addWidget(self.exit_btn)
        
        # Right panel with table
        right_panel = QFrame()
        right_panel.setFrameShape(QFrame.StyledPanel)
        right_layout = QVBoxLayout()
        right_panel.setLayout(right_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Инв. №", "Тип ОС", "Наименование", "Серийный №", "Статус", 
            "Подразделение", "МОЛ", "Местоположение", "Дата постановки", "Примечания"
        ])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.doubleClicked.connect(self.view_asset)
        self.table.selectionModel().selectionChanged.connect(self.update_ui_state)
        
        # Настройки для изменения размера и перетаскивания колонок
        header = self.table.horizontalHeader()
        # Разрешаем ручное изменение размера
        header.setSectionResizeMode(QHeaderView.Interactive)
        # Разрешаем перетаскивание колонок
        header.setSectionsMovable(True)
        # Разрешаем колонке "Наименование" растягиваться, но не делаем это автоматически
        header.setSectionResizeMode(2, QHeaderView.Interactive)
        
        # Устанавливаем начальные размеры колонок
        self.table.setColumnWidth(0, 100)  # Инв. №
        self.table.setColumnWidth(1, 120)  # Тип ОС
        self.table.setColumnWidth(2, 300)  # Наименование (шире по умолчанию)
        self.table.setColumnWidth(3, 120)  # Серийный №
        self.table.setColumnWidth(4, 100)  # Статус
        self.table.setColumnWidth(5, 150)  # Подразделение
        self.table.setColumnWidth(6, 100)  # МОЛ
        self.table.setColumnWidth(7, 150)  # Местоположение
        self.table.setColumnWidth(8, 120)  # Дата постановки
        self.table.setColumnWidth(9, 200)  # Колонка "Примечания"
        
        right_layout.addWidget(self.table)
        
        # Add panels to main layout
        main_layout.addWidget(left_panel)
        main_layout.addWidget(right_panel, 1)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # Устанавливаем начальные размеры колонок, но разрешаем их изменять
        self.table.horizontalHeader().resizeSection(0, 100)  # Инв. №
        self.table.horizontalHeader().resizeSection(1, 120)  # Тип ОС
        self.table.horizontalHeader().resizeSection(3, 120)  # Серийный №
        self.table.horizontalHeader().resizeSection(4, 100)  # Статус
        self.table.horizontalHeader().resizeSection(5, 150)  # Подразделение
        self.table.horizontalHeader().resizeSection(6, 100)  # МОЛ
        self.table.horizontalHeader().resizeSection(7, 150)  # Местоположение
        self.table.horizontalHeader().resizeSection(8, 120)  # Дата постановки
        # Update UI state
        self.update_ui_state()
        
        # Проверка подключения БД
        db_path = self.settings.value("database_path")
        if db_path and os.path.exists(db_path):
            try:
                self.db_manager = DatabaseManager(db_path)
                self.load_assets()
                QMessageBox.information(self, "Успех", f"База данных подключена: {db_path}")
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось подключиться к сохраненной базе данных: {str(e)}")
                self.settings.remove("database_path")
                self.show_database_dialog()
        else:
            self.show_database_dialog()
    
    def init_menu(self):
        menubar = self.menuBar()
        
        # Меню "Справочники"
        reference_menu = menubar.addMenu("Справочники")
        
        manage_asset_types_action = QAction("Управление типами ОС", self)
        manage_asset_types_action.triggered.connect(self.manage_asset_types)
        reference_menu.addAction(manage_asset_types_action)
        
        manage_departments_action = QAction("Управление подразделениями", self)
        manage_departments_action.triggered.connect(self.manage_departments)
        reference_menu.addAction(manage_departments_action)
        
        # Новое меню "Настройки"
        settings_menu = menubar.addMenu("Настройки")
        
        # Действия для смены темы
        light_theme_action = QAction("Светлая тема", self)
        light_theme_action.triggered.connect(lambda: self.set_theme("light"))
        settings_menu.addAction(light_theme_action)
        
        dark_theme_action = QAction("Темная тема", self)
        dark_theme_action.triggered.connect(lambda: self.set_theme("dark"))
        settings_menu.addAction(dark_theme_action)
        
        settings_menu.addSeparator()
        
        # Действие для смены БД
        change_db_action = QAction("Сменить базу данных...", self)
        change_db_action.triggered.connect(self.change_database)
        settings_menu.addAction(change_db_action)
    
    def manage_asset_types(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        dialog = ReferenceManagerDialog(self, 'asset_type')
        dialog.exec_()
        # Обновляем данные в основном интерфейсе, если нужно
        self.load_assets(self.current_filters)
    
    def manage_departments(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Ошибка", "База данных не подключена")
            return
            
        dialog = ReferenceManagerDialog(self, 'department')
        dialog.exec_()
        # Обновляем данные в основном интерфейсе, если нужно
        self.load_assets(self.current_filters)
    
    def update_ui_state(self):
        has_db = self.db_manager is not None
        has_selection = self.table.currentRow() >= 0
        
        self.add_btn.setEnabled(has_db)
        self.edit_btn.setEnabled(has_db and has_selection)
        self.delete_btn.setEnabled(has_db and has_selection)
        self.filter_btn.setEnabled(has_db)
        self.clear_filter_btn.setEnabled(has_db and self.current_filters is not None)
        self.print_btn.setEnabled(has_db and self.table.rowCount() > 0)
        self.export_btn.setEnabled(has_db and self.table.rowCount() > 0)
        
        # Показываем текущую базу данных в статусной строке
        if has_db:
            db_name = os.path.basename(self.db_manager.db_path)
            self.status_bar.showMessage(f"База: {db_name} | Записей: {self.table.rowCount()}")
        else:
            self.status_bar.showMessage("База данных не подключена")
    
    def show_database_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Подключение базы данных")
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.setFixedSize(400, 200)
        
        layout = QVBoxLayout()
        
        label = QLabel("Выберите действие с базой данных основных средств:")
        label.setWordWrap(True)
        layout.addWidget(label)
        
        create_btn = QPushButton("Создать новую базу данных")
        create_btn.clicked.connect(lambda: self.handle_db_action(dialog, 'create'))
        
        connect_btn = QPushButton("Подключить существующую базу данных")
        connect_btn.clicked.connect(lambda: self.handle_db_action(dialog, 'connect'))
        
        layout.addWidget(create_btn)
        layout.addWidget(connect_btn)
        dialog.setLayout(layout)
        
        dialog.exec_()
    
    def handle_db_action(self, dialog, action):
        if action == 'create':
            path, _ = QFileDialog.getSaveFileName(
                self, 
                "Создать базу данных", 
                "", 
                "SQLite Database (*.db);;All Files (*)"
            )
        
            if path:
                if not path.endswith('.db'):
                    path += '.db'
                try:
                    # Закрываем предыдущее соединение, если оно было
                    if self.db_manager:
                        self.db_manager.close()
                    
                    self.db_manager = DatabaseManager(path)
                    self.settings.setValue("database_path", path)
                    self.load_assets()
                    dialog.accept()
                    QMessageBox.information(self, "Успех", f"База данных успешно создана: {path}")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось создать базу данных: {str(e)}")
                    self.db_manager = None
    
        elif action == 'connect':
            path, _ = QFileDialog.getOpenFileName(
                self, 
                "Подключить базу данных", 
                "", 
                "SQLite Database (*.db);;All Files (*)"
            )
            
            if path:
                try:
                    # Закрываем предыдущее соединение, если оно было
                    if self.db_manager:
                        self.db_manager.close()
                    
                    self.db_manager = DatabaseManager(path)
                    self.settings.setValue("database_path", path)
                    self.load_assets()
                    dialog.accept()
                    QMessageBox.information(self, "Успех", f"База данных успешно подключена: {path}")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Не удалось подключиться к базе данных: {str(e)}")
                    self.db_manager = None
        
        self.update_ui_state()
    
    def load_assets(self, filters=None):
        if not self.db_manager:
            return
            
        self.current_filters = filters
        assets = self.db_manager.get_assets(filters)
        
        self.table.setRowCount(len(assets))
        
        for row, asset in enumerate(assets):
            self.table.setItem(row, 0, QTableWidgetItem(asset.get('inventory_number', '')))
            self.table.setItem(row, 1, QTableWidgetItem(asset.get('asset_type', '')))
            self.table.setItem(row, 2, QTableWidgetItem(asset.get('name', '')))
            self.table.setItem(row, 3, QTableWidgetItem(asset.get('serial_number', '')))
            self.table.setItem(row, 4, QTableWidgetItem(asset.get('status', '')))
            self.table.setItem(row, 5, QTableWidgetItem(asset.get('department', '')))
            self.table.setItem(row, 6, QTableWidgetItem(asset.get('mol', '')))
            self.table.setItem(row, 7, QTableWidgetItem(asset.get('location', '')))
            self.table.setItem(row, 8, QTableWidgetItem(asset.get('purchase_date', '')))
            self.table.setItem(row, 9, QTableWidgetItem(asset.get('notes', '')))
            
            # Store asset ID in the first column for easy access
            if 'id' in asset:
                self.table.item(row, 0).setData(Qt.UserRole, asset['id'])
        
        self.status_bar.showMessage(f"Найдено записей: {len(assets)}")
        self.update_ui_state()
    
    def add_asset(self):
        if not self.db_manager:
            return
            
        dialog = AssetDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            try:
                asset_data = dialog.get_data()
                self.db_manager.add_asset(asset_data)
                self.load_assets(self.current_filters)
                QMessageBox.information(self, "Успех", "Основное средство успешно добавлено")
            except ValueError as e:
                QMessageBox.warning(self, "Ошибка", str(e))
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить основное средство: {str(e)}")
    
    def edit_asset(self):
        if not self.db_manager:
            return
            
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return
            
        asset_id = self.table.item(selected_row, 0).data(Qt.UserRole)
        asset_data = self.db_manager.get_asset(asset_id)
        
        if not asset_data:
            return
            
        dialog = AssetDialog(self, asset_data, 'edit')
        if dialog.exec_() == QDialog.Accepted:
            try:
                new_data = dialog.get_data()
                self.db_manager.update_asset(asset_id, new_data)
                self.load_assets(self.current_filters)
                QMessageBox.information(self, "Успех", "Основное средство успешно обновлено")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось обновить основное средство: {str(e)}")
    
    def view_asset(self, index):
        try:
            if not self.db_manager:
                QMessageBox.warning(self, "Ошибка", "База данных не подключена")
                return
                
            selected_row = index.row()
            item = self.table.item(selected_row, 0)
            if not item:
                QMessageBox.warning(self, "Ошибка", "Не удалось выбрать запись")
                return
                
            asset_id = item.data(Qt.UserRole)
            if not asset_id:
                QMessageBox.warning(self, "Ошибка", "Не удалось получить ID основного средства")
                return
                
            asset_data = self.db_manager.get_asset(asset_id)
            if not asset_data:
                QMessageBox.warning(self, "Ошибка", "Основное средство не найдено в базе данных")
                return
                
            dialog = AssetDialog(self, asset_data, 'view')
            dialog.setWindowTitle(f"Просмотр ОС №{asset_data.get('inventory_number', '')}")
            
            # Делаем все поля только для чтения
            for child in dialog.findChildren(QLineEdit):
                child.setReadOnly(True)
            for child in dialog.findChildren(QComboBox):
                child.setEnabled(False)
                
            # Настраиваем кнопки (оставляем только Close)
            button_box = dialog.findChild(QDialogButtonBox)
            if button_box:
                button_box.clear()
                button_box.addButton(QDialogButtonBox.Close)
                
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при открытии карточки:\n{str(e)}")
    
    def delete_asset(self):
        if not self.db_manager:
            return
            
        selected_row = self.table.currentRow()
        if selected_row < 0:
            return
            
        asset_id = self.table.item(selected_row, 0).data(Qt.UserRole)
        asset_data = self.db_manager.get_asset(asset_id)
        
        if not asset_data:
            return
            
        reply = QMessageBox.question(
            self,
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить основное средство {asset_data['inventory_number']}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.delete_asset(asset_id)
                self.load_assets(self.current_filters)
                QMessageBox.information(self, "Успех", "Основное средство успешно удалено")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось удалить основное средство: {str(e)}")
    
    def show_filter_dialog(self):
        if not self.db_manager:
            return
            
        dialog = FilterDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            filters = dialog.get_filters()
            self.load_assets(filters)
    
    def clear_filters(self):
        self.load_assets()
    
    def print_assets(self):
        if not self.db_manager or self.table.rowCount() == 0:
            return
        
        printer = QPrinter(QPrinter.HighResolution)
        dialog = QPrintDialog(printer, self)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        painter = QPainter()
        if not painter.begin(printer):
            QMessageBox.warning(self, "Ошибка", "Не удалось начать печать")
            return
        
        try:
            # Получаем данные для печати
            headers = [self.table.horizontalHeaderItem(i).text() 
                      for i in range(self.table.columnCount())]
            rows = []
            for row in range(self.table.rowCount()):
                rows.append([self.table.item(row, col).text() 
                            for col in range(self.table.columnCount())])
            
            # Рисуем содержимое
            font = QFont("Arial", 10)
            painter.setFont(font)
            
            # Заголовки
            y = 100
            x = 100
            for i, header in enumerate(headers):
                painter.drawText(x, y, header)
                x += 150
            
            # Данные
            y += 30
            for row in rows:
                x = 100
                for i, item in enumerate(row):
                    painter.drawText(x, y, item)
                    x += 150
                y += 30
        finally:
            painter.end()
        # Получаем данные для печати
        headers = [self.table.horizontalHeaderItem(i).text() 
                  for i in range(self.table.columnCount())]
        rows = []
        for row in range(self.table.rowCount()):
            rows.append([self.table.item(row, col).text() 
                        for col in range(self.table.columnCount())])
        
        # Рисуем содержимое
        font = QFont("Arial", 10)
        painter.setFont(font)
        
        # Заголовки
        y = 100
        x = 100
        for i, header in enumerate(headers):
            painter.drawText(x, y, header)
            x += 150
        
        # Данные
        y += 30
        for row in rows:
            x = 100
            for i, item in enumerate(row):
                painter.drawText(x, y, item)
                x += 150
            y += 30
            
        painter.end()
    
    def export_to_excel(self):
        if not self.db_manager or self.table.rowCount() == 0:
            return
            
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт в Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not path:
            return
            
        if not path.endswith('.xlsx'):
            path += '.xlsx'
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Основные средства"
            
            # Write headers
            headers = [
                "Инв. №", "Тип ОС", "Наименование", "Серийный №", "Статус",
                "Подразделение", "МОЛ", "Местоположение", "Дата постановки", "Примечания"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            assets = self.db_manager.get_assets(self.current_filters)
            
            for row, asset in enumerate(assets, 2):
                ws.cell(row=row, column=1, value=asset.get('inventory_number', ''))
                ws.cell(row=row, column=2, value=asset.get('asset_type', ''))
                ws.cell(row=row, column=3, value=asset.get('name', ''))
                ws.cell(row=row, column=4, value=asset.get('serial_number', ''))
                ws.cell(row=row, column=5, value=asset.get('status', ''))
                ws.cell(row=row, column=6, value=asset.get('department', ''))
                ws.cell(row=row, column=7, value=asset.get('mol', ''))
                ws.cell(row=row, column=8, value=asset.get('location', ''))
                ws.cell(row=row, column=9, value=asset.get('purchase_date', ''))
                ws.cell(row=row, column=10, value=asset.get('notes', ''))
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(path)
            QMessageBox.information(self, "Успех", f"Данные успешно экспортированы в {path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать данные: {str(e)}")
    
    def show_context_menu(self, position):
        if not self.db_manager or self.table.currentRow() < 0:
            return
            
        menu = QMenu()
        
        history_action = QAction("История изменений", self)
        history_action.triggered.connect(self.show_asset_history)
        menu.addAction(history_action)
        
        view_action = QAction("Просмотр", self)
        view_action.triggered.connect(self.view_asset_from_context)
        menu.addAction(view_action)
        
        edit_action = QAction("Редактировать", self)
        edit_action.triggered.connect(self.edit_asset)
        menu.addAction(edit_action)
        
        delete_action = QAction("Удалить", self)
        delete_action.triggered.connect(self.delete_asset)
        menu.addAction(delete_action)
        
        menu.exec_(self.table.viewport().mapToGlobal(position))
        
    def show_asset_history(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0 and self.db_manager:
            asset_id = self.table.item(selected_row, 0).data(Qt.UserRole)
            dialog = HistoryDialog(self, asset_id, self.db_manager)
            dialog.exec_()
    
    def view_asset_from_context(self):
        index = self.table.currentIndex()
        self.view_asset(index)
    
    def set_theme(self, theme):
        if theme == "dark":
            # Dark theme
            palette = QPalette()
            palette.setColor(QPalette.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.WindowText, Qt.white)
            palette.setColor(QPalette.Base, QColor(25, 25, 25))
            palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ToolTipBase, Qt.white)
            palette.setColor(QPalette.ToolTipText, Qt.white)
            palette.setColor(QPalette.Text, Qt.white)
            palette.setColor(QPalette.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ButtonText, Qt.white)
            palette.setColor(QPalette.BrightText, Qt.red)
            palette.setColor(QPalette.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
            palette.setColor(QPalette.HighlightedText, Qt.black)
            
            self.setPalette(palette)
            self.settings.setValue("theme", "dark")
        else:
            # Light theme (default)
            self.setPalette(QApplication.style().standardPalette())
            self.settings.setValue("theme", "light")
            
    def change_database(self):
        """Позволяет пользователю выбрать другую базу данных"""
        reply = QMessageBox.question(
            self,
            "Смена базы данных",
            "Вы действительно хотите сменить базу данных? Текущие изменения будут сохранены.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        # Закрываем текущее соединение
        if self.db_manager:
            self.db_manager.close()
        
        # Показываем диалог выбора базы
        self.show_database_dialog()
        
        # Обновляем интерфейс
        self.update_ui_state()
        
    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            "Подтверждение выхода",
            "Вы уверены, что хотите выйти из программы?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.db_manager:
                self.db_manager.close()
            event.accept()
        else:
            event.ignore()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    
    # Set application font
    font = QFont()
    font.setFamily("Segoe UI")
    font.setPointSize(10)
    app.setFont(font)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())